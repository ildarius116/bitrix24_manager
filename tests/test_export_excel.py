"""Юнит-тесты build_workbook из src/export_excel.py (Фаза 2/5_04).

Без сети. Все xlsx записываются в tmp_path pytest (не в out/).
Синтетические WorkdayDay/WorkLog (как в test_workday.py).

Покрываемые единицы:
- build_workbook  — файл создаётся, содержит 3 листа с правильными именами
- summary        — main_rows / group_count / main_hours / group_hours / min_date / max_date
- Согласованность — main_hours == group_hours (суммы часов на обоих листах совпадают)
- Крайние случаи — дни без учётов, user_map разрешение, narrower-флаг
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.export_excel import MAIN_HEADERS, build_workbook
from src.workday import WorkdayDay, WorkLog


# ---------------------------------------------------------------------------
# Вспомогательные конструкторы
# ---------------------------------------------------------------------------

def _wl(
    desc: str,
    hours: float,
    title: str = "Общие задачи подразделения",
    wl_id: int = 0,
) -> WorkLog:
    """Создать WorkLog с заданным описанием и часами."""
    return WorkLog(
        id=wl_id,
        parent_day_id=None,
        title=title,
        description=desc,
        hours=hours,
        contract="T512_2",
        result=hours,
        raw={},
    )


def _day(
    d: date,
    logs: list,
    day_id: int = 1,
    employee: str = "1244",
    title: str = "Иванов И. И. | 01.06.2026",
) -> WorkdayDay:
    """Создать WorkdayDay с заданной датой и списком учётов."""
    return WorkdayDay(
        id=day_id,
        date=d,
        title=title,
        employee=employee,
        works_ids=[wl.id for wl in logs],
        raw={},
        logs=logs,
    )


# ---------------------------------------------------------------------------
# Фикстуры с синтетическими данными
# ---------------------------------------------------------------------------

DATE_A = date(2026, 6, 21)
DATE_B = date(2026, 6, 23)
DATE_C = date(2026, 6, 25)


def _sample_days() -> list:
    """Три дня с учётами для базовых тестов."""
    return [
        _day(DATE_A, [_wl("Задача Alpha", 4.0, wl_id=1), _wl("Задача Beta", 4.0, wl_id=2)], day_id=10),
        _day(DATE_B, [_wl("Задача Alpha", 8.0, wl_id=3)], day_id=20),
        _day(DATE_C, [_wl("Задача Gamma", 2.0, wl_id=4)], day_id=30),
    ]


# ---------------------------------------------------------------------------
# TestBuildWorkbookSheets
# ---------------------------------------------------------------------------

class TestBuildWorkbookSheets:
    """build_workbook создаёт xlsx с ровно тремя нужными листами."""

    def test_file_created(self, tmp_path):
        """Файл xlsx создаётся в указанном out_path."""
        out = tmp_path / "result.xlsx"
        days = _sample_days()
        build_workbook(days, DATE_A, DATE_C, out)
        assert out.exists()

    def test_three_sheets(self, tmp_path):
        """Рабочая книга содержит ровно 3 листа."""
        out = tmp_path / "result.xlsx"
        build_workbook(_sample_days(), DATE_A, DATE_C, out)
        wb = load_workbook(out)
        assert len(wb.sheetnames) == 3

    def test_sheet_names(self, tmp_path):
        """Имена листов: 'Рабочий день', 'Группировка', 'Метаданные'."""
        out = tmp_path / "result.xlsx"
        build_workbook(_sample_days(), DATE_A, DATE_C, out)
        wb = load_workbook(out)
        assert wb.sheetnames == ["Рабочий день", "Группировка", "Метаданные"]

    def test_main_sheet_header_row(self, tmp_path):
        """Первая строка основного листа — заголовки."""
        out = tmp_path / "result.xlsx"
        build_workbook(_sample_days(), DATE_A, DATE_C, out)
        wb = load_workbook(out)
        ws = wb["Рабочий день"]
        header = [ws.cell(1, col + 1).value for col in range(len(MAIN_HEADERS))]
        assert header == MAIN_HEADERS


# ---------------------------------------------------------------------------
# TestBuildWorkbookSummary
# ---------------------------------------------------------------------------

class TestBuildWorkbookSummary:
    """build_workbook возвращает корректный словарь сводки."""

    def test_main_rows_count(self, tmp_path):
        """main_rows = суммарное число пар (день × учёт)."""
        out = tmp_path / "r.xlsx"
        # 2 учёта у первого + 1 у второго + 1 у третьего = 4 строки данных
        days = _sample_days()
        summary = build_workbook(days, DATE_A, DATE_C, out)
        assert summary["main_rows"] == 4

    def test_group_count(self, tmp_path):
        """group_count = число уникальных «Описаний задачи»."""
        out = tmp_path / "r.xlsx"
        days = _sample_days()
        summary = build_workbook(days, DATE_A, DATE_C, out)
        # Alpha, Beta, Gamma → 3 группы
        assert summary["group_count"] == 3

    def test_min_date(self, tmp_path):
        """min_date = минимальная фактическая дата."""
        out = tmp_path / "r.xlsx"
        summary = build_workbook(_sample_days(), DATE_A, DATE_C, out)
        assert summary["min_date"] == DATE_A

    def test_max_date(self, tmp_path):
        """max_date = максимальная фактическая дата."""
        out = tmp_path / "r.xlsx"
        summary = build_workbook(_sample_days(), DATE_A, DATE_C, out)
        assert summary["max_date"] == DATE_C

    def test_main_hours_correct(self, tmp_path):
        """main_hours = сумма часов всех учётов."""
        out = tmp_path / "r.xlsx"
        days = _sample_days()
        # 4 + 4 + 8 + 2 = 18 часов
        summary = build_workbook(days, DATE_A, DATE_C, out)
        assert summary["main_hours"] == pytest.approx(18.0)

    def test_group_hours_correct(self, tmp_path):
        """group_hours = сумма часов по листу Группировка."""
        out = tmp_path / "r.xlsx"
        days = _sample_days()
        summary = build_workbook(days, DATE_A, DATE_C, out)
        assert summary["group_hours"] == pytest.approx(18.0)

    def test_path_in_summary(self, tmp_path):
        """summary['path'] содержит путь к созданному файлу."""
        out = tmp_path / "r.xlsx"
        summary = build_workbook(_sample_days(), DATE_A, DATE_C, out)
        assert summary["path"] == str(out)


# ---------------------------------------------------------------------------
# TestBuildWorkbookHoursConsistency
# ---------------------------------------------------------------------------

class TestBuildWorkbookHoursConsistency:
    """Сумма часов основного листа == сумме часов листа группировки."""

    def test_hours_main_eq_group_simple(self, tmp_path):
        """Базовый кейс: main_hours == group_hours."""
        out = tmp_path / "r.xlsx"
        summary = build_workbook(_sample_days(), DATE_A, DATE_C, out)
        assert summary["main_hours"] == pytest.approx(summary["group_hours"])

    def test_hours_consistency_multiple_logs_per_day(self, tmp_path):
        """Несколько учётов у одного дня: суммы совпадают."""
        out = tmp_path / "r.xlsx"
        days = [
            _day(DATE_A, [_wl("A", 2.0), _wl("A", 3.0), _wl("B", 5.0)], day_id=1),
        ]
        summary = build_workbook(days, DATE_A, DATE_C, out)
        assert summary["main_hours"] == pytest.approx(summary["group_hours"])
        assert summary["main_hours"] == pytest.approx(10.0)

    def test_hours_consistency_no_logs(self, tmp_path):
        """Дни без учётов: обе суммы = 0."""
        out = tmp_path / "r.xlsx"
        days = [
            _day(DATE_A, [], day_id=1),
            _day(DATE_B, [], day_id=2),
        ]
        summary = build_workbook(days, DATE_A, DATE_C, out)
        assert summary["main_hours"] == 0.0
        assert summary["group_hours"] == 0.0

    def test_hours_consistency_none_hours(self, tmp_path):
        """WorkLog с hours=None трактуется как 0 — суммы совпадают."""
        out = tmp_path / "r.xlsx"
        wl_none = WorkLog(id=9, parent_day_id=None, title="t", description="X",
                          hours=None, contract="", result=None, raw={})
        wl_ok = WorkLog(id=10, parent_day_id=None, title="t", description="Y",
                        hours=3.0, contract="", result=None, raw={})
        days = [_day(DATE_A, [wl_none, wl_ok], day_id=1)]
        summary = build_workbook(days, DATE_A, DATE_C, out)
        assert summary["main_hours"] == pytest.approx(summary["group_hours"])
        assert summary["main_hours"] == pytest.approx(3.0)


# ---------------------------------------------------------------------------
# TestBuildWorkbookNoDays
# ---------------------------------------------------------------------------

class TestBuildWorkbookNoDays:
    """build_workbook с пустым списком и днями без учётов."""

    def test_empty_days_list(self, tmp_path):
        """Пустой список дней — файл создаётся, main_rows=0."""
        out = tmp_path / "r.xlsx"
        summary = build_workbook([], DATE_A, DATE_C, out)
        assert out.exists()
        assert summary["main_rows"] == 0
        assert summary["group_count"] == 0
        assert summary["main_hours"] == 0.0
        assert summary["min_date"] is None
        assert summary["max_date"] is None

    def test_day_without_logs_one_empty_row(self, tmp_path):
        """День без учётов добавляет одну строку (пустые поля учёта)."""
        out = tmp_path / "r.xlsx"
        days = [_day(DATE_A, [], day_id=1)]
        summary = build_workbook(days, DATE_A, DATE_C, out)
        # Одна строка (день без учётов → 1 пустая строка по учёту)
        assert summary["main_rows"] == 1
        assert summary["group_count"] == 0  # нет учётов → нет групп


# ---------------------------------------------------------------------------
# TestBuildWorkbookNarrower
# ---------------------------------------------------------------------------

class TestBuildWorkbookNarrower:
    """Флаг narrower: фактические рамки уже заданного периода."""

    def test_no_narrower_when_data_fits_period(self, tmp_path):
        """Данные покрывают весь период → narrower=False."""
        out = tmp_path / "r.xlsx"
        # date_from=DATE_A, date_to=DATE_C; данные [DATE_A..DATE_C]
        summary = build_workbook(_sample_days(), DATE_A, DATE_C, out)
        assert summary["narrower"] is False

    def test_narrower_when_data_shorter_than_period(self, tmp_path):
        """Данные покрывают только часть периода → narrower=True."""
        out = tmp_path / "r.xlsx"
        # period: 2026-06-01..2026-06-30, data: DATE_A..DATE_C (уже)
        wide_from = date(2026, 6, 1)
        wide_to = date(2026, 6, 30)
        summary = build_workbook(_sample_days(), wide_from, wide_to, out)
        assert summary["narrower"] is True

    def test_empty_days_no_narrower(self, tmp_path):
        """Пустой список дней → narrower=False (нет данных для сравнения)."""
        out = tmp_path / "r.xlsx"
        summary = build_workbook([], DATE_A, DATE_C, out)
        assert summary["narrower"] is False


# ---------------------------------------------------------------------------
# TestBuildWorkbookUserMap
# ---------------------------------------------------------------------------

class TestBuildWorkbookUserMap:
    """user_map: разрешение user_id → ФИО."""

    def test_user_map_resolves_employee(self, tmp_path):
        """Если user_map задан — колонка 'Сотрудник' содержит ФИО."""
        out = tmp_path / "r.xlsx"
        days = [_day(DATE_A, [_wl("A", 8.0)], day_id=1, employee="1244")]
        user_map = {"1244": "Иванов Иван Иванович"}
        build_workbook(days, DATE_A, DATE_C, out, user_map=user_map)
        wb = load_workbook(out)
        ws = wb["Рабочий день"]
        # Строка 2 — первая строка данных; колонка 3 — «Сотрудник»
        employee_cell = ws.cell(2, 3).value
        assert employee_cell == "Иванов Иван Иванович"

    def test_no_user_map_shows_raw_id(self, tmp_path):
        """Если user_map не задан — колонка 'Сотрудник' содержит сырой id."""
        out = tmp_path / "r.xlsx"
        days = [_day(DATE_A, [_wl("A", 8.0)], day_id=1, employee="9999")]
        build_workbook(days, DATE_A, DATE_C, out)
        wb = load_workbook(out)
        ws = wb["Рабочий день"]
        employee_cell = ws.cell(2, 3).value
        assert employee_cell == "9999"

    def test_missing_user_in_map_fallback_raw(self, tmp_path):
        """user_id не в user_map → выводится сырой id (fallback)."""
        out = tmp_path / "r.xlsx"
        days = [_day(DATE_A, [_wl("A", 8.0)], day_id=1, employee="7777")]
        user_map = {"1244": "Другой сотрудник"}  # 7777 отсутствует
        build_workbook(days, DATE_A, DATE_C, out, user_map=user_map)
        wb = load_workbook(out)
        ws = wb["Рабочий день"]
        employee_cell = ws.cell(2, 3).value
        assert employee_cell == "7777"


# ---------------------------------------------------------------------------
# TestBuildWorkbookDateRange
# ---------------------------------------------------------------------------

class TestBuildWorkbookDateRange:
    """Фактические min/max дат из данных (игнорируя None-даты)."""

    def test_single_day_min_eq_max(self, tmp_path):
        """Один день → min_date == max_date."""
        out = tmp_path / "r.xlsx"
        days = [_day(DATE_B, [_wl("A", 8.0)], day_id=1)]
        summary = build_workbook(days, DATE_A, DATE_C, out)
        assert summary["min_date"] == DATE_B
        assert summary["max_date"] == DATE_B

    def test_day_without_date_ignored_in_range(self, tmp_path):
        """День с date=None не учитывается в min/max."""
        out = tmp_path / "r.xlsx"
        days = [
            _day(None, [_wl("A", 4.0)], day_id=1),  # date=None
            _day(DATE_C, [_wl("B", 4.0)], day_id=2),
        ]
        summary = build_workbook(days, DATE_A, DATE_C, out)
        assert summary["min_date"] == DATE_C
        assert summary["max_date"] == DATE_C
