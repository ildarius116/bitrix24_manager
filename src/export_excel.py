"""Формирование Excel-выгрузки «Рабочий день» (openpyxl).

Фаза 2 (FR-1). Три листа:
- «Рабочий день» — одна строка = день × его учёт (день без учётов → одна пустая по учёту);
- «Группировка» — агрегаты по «Описанию задачи» (FR-1.2);
- «Метаданные» — фактические рамки дат и сверка с заданным периодом (FR-1.1.6).

Только формирование файла (никаких сетевых вызовов). Данные приходят из src.workday.
"""

from __future__ import annotations

import logging
from collections import OrderedDict
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .workday import WorkdayDay, WorkLog, data_time_range

# Тип: словарь {employee_id_str: ФИО}.
UserMap = Dict[str, str]

log = logging.getLogger("workday")

_HEADER_FONT = Font(bold=True)
_DATE_FORMAT = "DD.MM.YYYY"

# Колонки основного листа.
# «Договор» убрана по запросу заказчика (2026-06-25); поле log_contract остаётся в Config.
MAIN_HEADERS = [
    "Дата",
    "Название",
    "Сотрудник",
    "Описание задачи",
    "Информация о проделанной работе",
    "Часы",
]

GROUP_HEADERS = [
    "Описание задачи",
    "Кол-во записей",
    "Сумма часов",
    "Даты",
]

META_HEADERS = ["Показатель", "Значение"]

NO_DESCRIPTION = "(без описания)"


def _fmt_date(d: Optional[date]) -> str:
    return d.strftime("%d.%m.%Y") if d else ""


def _style_header(ws: Worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="top")


def _autosize(ws: Worksheet, ncols: int, *, max_width: int = 60) -> None:
    """Грубая авто-ширина по содержимому (ограничена max_width)."""
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        width = 0
        for cell in ws[letter]:
            value = cell.value
            if value is None:
                continue
            length = max((len(line) for line in str(value).splitlines()), default=0)
            width = max(width, length)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def _resolve_employee(employee_raw: str, user_map: "UserMap") -> str:
    """Вернуть ФИО по сырому значению поля employee.

    Пробует преобразовать значение к int (user id) и найти в user_map.
    Если не найдено или не парсится — возвращает исходную строку (fallback).
    """
    if not employee_raw:
        return employee_raw
    try:
        uid = int(employee_raw)
    except (TypeError, ValueError):
        return employee_raw
    return user_map.get(str(uid), employee_raw)


def _write_main_sheet(
    ws: Worksheet,
    days: List[WorkdayDay],
    user_map: "Optional[UserMap]" = None,
) -> int:
    """Заполнить основной лист. Возвращает число строк данных (без заголовка).

    user_map — словарь {str(user_id): ФИО}, если передан — используется для
    колонки «Сотрудник». Если None или id не найден — выводится сырое значение.
    """
    ws.title = "Рабочий день"
    ws.append(MAIN_HEADERS)
    _style_header(ws, len(MAIN_HEADERS))

    umap: UserMap = user_map or {}

    row_count = 0
    for day in days:
        employee_display = _resolve_employee(day.employee, umap)
        logs: List[WorkLog] = day.logs or []
        if not logs:
            # День без учётов — одна строка с пустыми полями учёта.
            ws.append([day.date, day.title, employee_display, "", "", None])
            row_count += 1
        else:
            for wl in logs:
                # «Описание задачи» — имя задачи из title учёта («Общие задачи
                # подразделения»); «Информация о проделанной работе» — текст из поля
                # «Описание задачи» (ufCrm48_1744239302). См. разбор в src/workday.py.
                ws.append(
                    [
                        day.date,
                        day.title,
                        employee_display,
                        wl.title,
                        wl.work_info,
                        wl.hours,
                    ]
                )
                row_count += 1

    # Формат даты в первой колонке (строки данных).
    for r in range(2, row_count + 2):
        ws.cell(row=r, column=1).number_format = _DATE_FORMAT

    _autosize(ws, len(MAIN_HEADERS))
    return row_count


def _group_rows(days: List[WorkdayDay]) -> List[Tuple[str, int, float, List[date]]]:
    """Сгруппировать учёты по «Описанию задачи».

    Ключ — описание (пустое → «(без описания)»). Агрегаты: кол-во, сумма часов, даты.
    Сортировка по описанию (регистронезависимо). Сумма часов считается по тем же
    значениям, что и на основном листе.
    """
    groups: "OrderedDict[str, Dict]" = OrderedDict()
    for day in days:
        for wl in day.logs or []:
            key = wl.description.strip() or NO_DESCRIPTION
            g = groups.get(key)
            if g is None:
                g = {"count": 0, "hours": 0.0, "dates": []}
                groups[key] = g
            g["count"] += 1
            g["hours"] += wl.hours or 0.0
            if day.date is not None and day.date not in g["dates"]:
                g["dates"].append(day.date)

    rows: List[Tuple[str, int, float, List[date]]] = []
    for key, g in groups.items():
        rows.append((key, g["count"], g["hours"], sorted(g["dates"])))
    rows.sort(key=lambda r: r[0].lower())
    return rows


def _write_group_sheet(ws: Worksheet, days: List[WorkdayDay]) -> Tuple[int, float]:
    """Заполнить лист «Группировка». Возвращает (число групп, общая сумма часов)."""
    ws.title = "Группировка"
    ws.append(GROUP_HEADERS)
    _style_header(ws, len(GROUP_HEADERS))

    rows = _group_rows(days)
    total_hours = 0.0
    for key, count, hours, dates in rows:
        dates_str = ", ".join(_fmt_date(d) for d in dates)
        ws.append([key, count, hours, dates_str])
        total_hours += hours

    _autosize(ws, len(GROUP_HEADERS))
    return len(rows), total_hours


def _write_meta_sheet(
    ws: Worksheet,
    days: List[WorkdayDay],
    main_rows: int,
    date_from: date,
    date_to: date,
) -> Tuple[bool, Optional[date], Optional[date]]:
    """Заполнить лист «Метаданные».

    Возвращает (narrower, min_date, max_date), где narrower=True, если фактические
    рамки данных уже заданного периода (повод для предупреждения).
    """
    ws.title = "Метаданные"
    ws.append(META_HEADERS)
    _style_header(ws, len(META_HEADERS))

    min_d, max_d, dated_count = data_time_range(days)
    narrower = False
    if min_d is not None and max_d is not None:
        narrower = (min_d > date_from) or (max_d < date_to)

    rows = [
        ("Всего строк (день × учёт)", main_rows),
        ("Всего дней (1208)", len(days)),
        ("Дней с распознанной датой", dated_count),
        ("Дата min (факт)", _fmt_date(min_d)),
        ("Дата max (факт)", _fmt_date(max_d)),
        ("Заданный период: с", _fmt_date(date_from)),
        ("Заданный период: по", _fmt_date(date_to)),
        ("Рамки уже периода", "да" if narrower else "нет"),
    ]
    for name, value in rows:
        ws.append([name, value])

    _autosize(ws, len(META_HEADERS))
    return narrower, min_d, max_d


def build_workbook(
    days: List[WorkdayDay],
    date_from: date,
    date_to: date,
    out_path: Path,
    *,
    user_map: Optional["UserMap"] = None,
) -> Dict[str, object]:
    """Собрать xlsx с тремя листами и сохранить в out_path.

    user_map — словарь {str(user_id): ФИО} для колонки «Сотрудник». Если None —
    выводится сырой id (обратная совместимость).

    Возвращает сводку: число строк основного листа, число групп, сумма часов (основной
    лист и группировка — должны совпадать), фактические min/max дат, флаг сужения рамок.
    """
    wb = Workbook()
    ws_main = wb.active  # первый лист
    main_rows = _write_main_sheet(ws_main, days, user_map=user_map)

    ws_group = wb.create_sheet("Группировка")
    group_count, group_hours = _write_group_sheet(ws_group, days)

    ws_meta = wb.create_sheet("Метаданные")
    narrower, min_d, max_d = _write_meta_sheet(ws_meta, days, main_rows, date_from, date_to)

    # Сумма часов по основному листу (для сверки с группировкой).
    main_hours = 0.0
    for day in days:
        for wl in day.logs or []:
            main_hours += wl.hours or 0.0

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    if narrower:
        log.warning(
            "Фактические рамки данных (%s … %s) уже заданного периода (%s … %s): "
            "возможна обрезка пагинацией или отсутствие данных на краях периода.",
            _fmt_date(min_d),
            _fmt_date(max_d),
            _fmt_date(date_from),
            _fmt_date(date_to),
        )

    return {
        "main_rows": main_rows,
        "group_count": group_count,
        "group_hours": round(group_hours, 6),
        "main_hours": round(main_hours, 6),
        "min_date": min_d,
        "max_date": max_d,
        "narrower": narrower,
        "path": str(out_path),
    }
